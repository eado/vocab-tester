from openpyxl import load_workbook

word_list = [
    {'word': "A", 'pinyin': "a", 'stage': 0, 'days': 0}, 
    {'word': "B", 'pinyin': "b", 'stage': 0, 'days': 0}, 
    {'word': "C", 'pinyin': "c", 'stage': 0, 'days': 0}, 
            ]
test = []

#config
stages = [1, 1, 2, 3, 5, 7]
test_filename = 'test.txt' #not used yet
vocab_filename = 'vocab.txt'
vocab_sheet_filename = 'HSK1.xlsx'
vocab_format = ['word', 'pinyin', 'stage', 'days']
vocab_column_seperator = '#'


#adds all rows in the word list with days < 1 to test
def build_test():
    del test[:] # clears the table test, as opposed to creating a NEW local table test = []
    for row in word_list: 
        if int(row['days']) < 1:
            test.append(row)

def grade_test():
    for row in word_list:
        if row in test:
            
                
            if input("{} {}: ".format(row['word'], row['pinyin'])):
                row['stage'] = 0
            else:
                row['stage'] = int(row['stage']) + 1
            row['days'] = stages[row['stage']]
                
def new_day():
    for row in word_list:
        row['days'] = row['days'] - 1
    
def write_test():
    with open(test_filename, 'w') as test_file:
        test_file.write("")
        for row in test:
            test_file.write("{}  ___\n".format(row['pinyin']))
                   
def load_word_list():
    del word_list[:] #clears word list
    with open(vocab_filename, 'r') as vocab_file: #opens the vocab file
        for file_row in vocab_file.readlines(): #populates word list with rows
            new_row = {} 
            for column in vocab_format: #populates each row with columns
                new_row[column] = file_row.split(vocab_column_seperator)[vocab_format.index(column)] #splits the rows of the vocab file, assigns the values to the dictionary new_row
            word_list.append(new_row)   
            
def load_word_list_from_sheet():
    del word_list[:]
    wb = load_workbook(vocab_sheet_filename)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0]) #sets sheet to the first sheet in the workbook
    
    line_number = 1
    while True:       
        new_row = {}
        column_number = 1
        while True:
            current_cell = sheet.cell(row=line_number, column=column_number)
            
            if not current_cell.value == None: #if the current cell is not None (if it is not empty)
                new_row[vocab_format[column_number - 1]] = current_cell.value #the -1 is because 
                #vocab_format is 0-indexed and the columns of the sheet are 1-indexed. 
            else:
                break
            column_number = column_number + 1
            
        if new_row: #if the new row is not empty
            word_list.append(new_row)
        else:
            break
            
        line_number = line_number + 1
            
            
i = 0        
load_word_list_from_sheet()
while True:   
    print("Day {}".format(i))
    build_test()
    write_test()
    grade_test()
    new_day()
    i = i + 1
    print("\n\n")
